#!/usr/bin/env python
import os, sys
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'music_registry.settings')
sys.path.insert(0, '/home/jav/SGA1')
import django
django.setup()

import openpyxl
from users.models import Usuario
from students.models import Student
from classes.models import GradeLevel

COURSE_LEVEL = {
    '1o': '1', '2o': '2', '3o': '3', '4o': '4', '5o': '5',
    '6o': '6', '7o': '7', '8o': '8', '9o ': '9', '10o': '10',
    '11o': '11', '1o Extra': '1'
}

excel_file = '/home/jav/SGA1/base_de_datos_json/25-26 Matriculados Conservatorio Bolívar de AMbato2.xlsx'
wb = openpyxl.load_workbook(excel_file)

seen_emails = set(u.email for u in Usuario.objects.exclude(email__isnull=True))
total_created = 0

for sheet_name in sorted(wb.sheetnames):
    if sheet_name not in COURSE_LEVEL:
        continue

    ws = wb[sheet_name]
    level = COURSE_LEVEL[sheet_name]

    # Get headers
    header = [cell.value for cell in ws[1]]
    cedula_col = next((i for i, h in enumerate(header) if h and 'Cédula' in str(h)), None)
    apellido_col = next((i for i, h in enumerate(header) if h and 'Apellid' in str(h)), None)
    nombre_col = next((i for i, h in enumerate(header) if h and 'Nombre' in str(h) and 'apellid' not in str(h).lower()), None)
    email_col = next((i for i, h in enumerate(header) if h and 'correo' in str(h).lower()), None)

    created = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        cedula = str(row[cedula_col]).strip() if cedula_col is not None and row[cedula_col] else None
        if not cedula or cedula == 'None' or len(cedula) < 3:
            continue

        # Skip if exists
        if Usuario.objects.filter(cedula=cedula).exists():
            continue

        apellido = str(row[apellido_col]).strip() if apellido_col is not None and row[apellido_col] else ''
        nombre = str(row[nombre_col]).strip() if nombre_col is not None and row[nombre_col] else ''
        email = str(row[email_col]).strip() if email_col is not None and row[email_col] else ''

        full_name = f"{apellido} {nombre}".strip() if (apellido or nombre) else cedula
        final_email = email if (email and '@' in email and email not in seen_emails) else None

        if final_email:
            seen_emails.add(final_email)

        try:
            usuario = Usuario.objects.create(
                rol=Usuario.Rol.ESTUDIANTE,
                cedula=cedula,
                nombre=full_name,
                email=final_email
            )

            gl, _ = GradeLevel.objects.get_or_create(level=level, section='')
            Student.objects.get_or_create(usuario=usuario, defaults={'grade_level': gl, 'active': True})

            created += 1
            total_created += 1
        except Exception as e:
            pass

    if created > 0:
        print(f"✓ {sheet_name:10} ({level}): {created} new")

print(f"\n{'='*50}")
print(f"Created: {total_created}")
print(f"Total: {Usuario.objects.filter(rol=Usuario.Rol.ESTUDIANTE).count()}")
